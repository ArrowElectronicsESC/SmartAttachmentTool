import pandas as pd
import datetime
import sys
import os
from unicodedata import category
import openpyxl

if len(sys.argv) == 3:
    IccMainPart = str(sys.argv[1])
    IccProperty = str(sys.argv[2])
    e = datetime.datetime.now()
    nameNewFile = IccMainPart + IccProperty + e.strftime("%Y%m%d%H%M%S") + ".xlsx"
    auxFile = "aux1.xlsx"
    pathfile = 'UnityReport.xlsx'

    writer = pd.ExcelWriter(nameNewFile, engine='xlsxwriter')
    writer.save()
    writer1 = pd.ExcelWriter(auxFile, engine='xlsxwriter')
    writer1.save()

    columns = ['partyName','LocationID','fsrName','Region','edRegion','salesLocation','dwIndicator','projectName','boardName',
               'boardsYr1','Prototype Date','Production Date','AUN','regstatus','regNum','faeName','supplier','supplierDivision','supplierPartNumber',
               'dwMargin','submittedDate','vendorApprovedDate','vendorExpirationDate','designWinDate','regProjectedYr1Rev','regNSBtoDate','fiscalMonth',
               'fiscalQtr','fiscalYear','metric','Note','topOpportunityFlag','DWB','BSV','icc2Name','icc3Name']

    df = pd.read_excel(pathfile)
    xl = pd.ExcelFile(pathfile)
    df.sort_values(by='Prototype Date')
    projectName_change = df["projectName"].shift() != df["projectName"]
    filesheet = xl.sheet_names
    infofile = xl.parse(filesheet[0])
    totalRows = infofile.shape[0] - 1
    testing = df.iat[0,7]
    columnIcc = 35 #icc3Name
    index = 0
    while index <= totalRows:
        if ((totalRows == index) and (projectName_change.iloc[index] == True)):
            if(df.iat[index,columnIcc] == IccMainPart):
                dataB1=df.loc[[index]]
                iccRowFound = pd.DataFrame(dataB1)
                df1=pd.read_excel(nameNewFile)
                df3=pd.concat([df1,iccRowFound])
                df3.to_excel(nameNewFile,index=False)
            index += 1
        elif((projectName_change.iloc[index] == True) and (projectName_change.iloc[index+1] == True)):
            if(df.iat[index,columnIcc] == IccMainPart):
                dataB1=df.loc[[index]]
                iccRowFound = pd.DataFrame(dataB1)
                df1=pd.read_excel(nameNewFile)
                df3=pd.concat([df1,iccRowFound])
                df3.to_excel(nameNewFile,index=False)
            index += 1
        elif ((projectName_change.iloc[index] == True) and (projectName_change.iloc[index+1] == False)):
            flagIccMainPart = 0
            flagIccProperty = 0
            auxCounter = 1
            if(df.iat[index,columnIcc] == IccMainPart):
                dataB1=df.loc[[index]]
                iccRowFound = pd.DataFrame(dataB1)
                df1=pd.read_excel(auxFile)
                df3=pd.concat([df1,iccRowFound])
                df3.to_excel(auxFile,index=False)
                flagIccMainPart = 1
            if(df.iat[index,columnIcc] == IccProperty):
                flagIccProperty = 1
            while (projectName_change.iloc[index+auxCounter] == False):
                if(df.iat[index+auxCounter,columnIcc] == IccMainPart):
                    dataB1=df.loc[[index+auxCounter]]
                    iccRowFound = pd.DataFrame(dataB1)
                    df1=pd.read_excel(auxFile)
                    df3=pd.concat([df1,iccRowFound])
                    df3.to_excel(auxFile,index=False)
                    flagIccMainPart = 1
                if(df.iat[index,columnIcc] == IccProperty):
                    flagIccProperty = 1
                auxCounter += 1

            if ((flagIccMainPart == 1) and (flagIccProperty == 0)):
                df1=pd.read_excel(auxFile)
                df2=pd.read_excel(nameNewFile)
                df3=pd.concat([df1,df2])
                df3.to_excel(nameNewFile,index=False)

            if os.path.exists(auxFile):
                writer1.handles.close()
                os.remove(auxFile)
                writer1 = pd.ExcelWriter(auxFile, engine='xlsxwriter')
                writer1.save()
            index = index + auxCounter - 1
        else:
            #do nothing
            index += 1
    writer1.handles.close()
    os.remove(auxFile)

    def lowercase_and_remove_non_alphanumeric(s):
        return ''.join(c for c in s if c.isalnum()).lower()

    pxl_doc = openpyxl.load_workbook("./Antennas Selector Guide.xlsx")
    sheet = pxl_doc['Bluetooth']

    antennasInformation=[]
    antennasName2Information = {}
    headers = []

    for i in range(1, sheet.max_column+1):
        title = sheet.cell(row=1, column=i).value
        headers.append(title)
    for i in range(2, sheet.max_row+1):
        antenna = {}
        for j in range(len(headers)):
            value = sheet.cell(row=i, column=j+1).value
            if value is not None and not (type(value) == str and value[0] is '='):
                antenna[headers[j]] = value
            else:
                antenna[headers[j]] = antennasInformation[-1][headers[j]]
            # antenna[headers[j]] = value
        antennasInformation.append(antenna)
        antennasName2Information[antenna['Part Number']] = antenna

    def findAntennasInFrequencyRange(minFreq, maxFreq):
        antennas = []
        for antenna in antennasInformation:
            if float(antenna['Frequency Min (MHz)']) <= minFreq <= float(antenna['Frequency Max (MHz)']) or float(antenna['Frequency Min (MHz)']) <= maxFreq <= float(antenna['Frequency Max (MHz)']):
                antennas.append(antenna["Part Number"])
        return antennas

    pxl_doc = openpyxl.load_workbook("./Bluetooth Selector Guide.xlsx")
    sheet = pxl_doc['Chip Level']

    chipLevelParts = set([])
    bluetoothInformation = {}
    for i in range(2, sheet.max_row+1):
        bt = sheet.cell(row=i, column=3).value
        if bt is not None:
            chipLevelParts.add(lowercase_and_remove_non_alphanumeric(str(bt)))
            bluetoothInformation[lowercase_and_remove_non_alphanumeric(str(bt))] = (float(sheet.cell(row=i, column=12).value), float(sheet.cell(row=i, column=13).value))

    ModuleParts = set([])
    bluetoothModules = {}
    sheet = pxl_doc['Modules']
    for i in range(1, sheet.max_row+1):
        bt = sheet.cell(row=i, column=3).value
        if bt is not None:
            ModuleParts.add(lowercase_and_remove_non_alphanumeric(str(bt)))

    def splitAntennasByMountingType(antennas):
        antennasByType = {}
        for antenna in antennas:
            # print(antenna)
            if antennasName2Information[antenna]['Mounting'] not in antennasByType:
                antennasByType[antennasName2Information[antenna]['Mounting']] = []
            antennasByType[antennasName2Information[antenna]['Mounting']].append(antenna)
        return antennasByType

    def p2f(x):
        return float(x.strip('%'))/100

    def sortAntennasByEfficiency(antennasByType):
        for mountingType in antennasByType:
            antennasByType[mountingType] = sorted(antennasByType[mountingType], key=lambda k: antennasName2Information[k]['Average Efficiency'], reverse=True)
        return antennasByType

    def findCompatibleParts(bluetooth):
        if bluetooth in chipLevelParts:
            antennas = []
            bluetoothFrequency = bluetoothInformation[bluetooth]
            inRange = findAntennasInFrequencyRange(bluetoothFrequency[0], bluetoothFrequency[1])
            antennas = list(set(antennas + inRange))
            for antenna in antennas:
                if antenna not in antennasName2Information:
                    antennas.remove(antenna)

            byType = splitAntennasByMountingType(antennas)
            sortedAntennas = sortAntennasByEfficiency(byType)

            adhesiveAntennas = sortedAntennas["Adhesive"][:3]
            chasisAntennas = sortedAntennas["Chasis Mount"][:3]
            panelAntennas = sortedAntennas["Panel Mount"][:3]
            surfaceAntennas = sortedAntennas["Surface Mount"][:3]

            return {'Adhesive': adhesiveAntennas, 'Chasis Mount': chasisAntennas, 'Panel Mount': panelAntennas, 'Surface Mount': surfaceAntennas}
        else:
            return {'Adhesive': [], 'Chasis Mount': [], 'Panel Mount': [], 'Surface Mount': []}

    parts2evaluate = []

    pxl_doc = openpyxl.load_workbook(nameNewFile)
    sheet = pxl_doc['Sheet1']


    # create directory named output_excel if it doesn't exist
    os.makedirs("./output_excel", exist_ok=True)

    bluetooth = None
    for i in range(2, sheet.max_row+1):
        bt = sheet.cell(row=i, column=19).value
        if bt is not None:
            bluetooth = lowercase_and_remove_non_alphanumeric(bt)
            parts = findCompatibleParts(bluetooth)
            with open(f"./output_excel/{bluetooth}_compatible_parts.csv", "w") as f:
                f.write("Adhesive,Chasis Mount,Panel Mount,Surface Mount,Modules\n")
                for i in range(3):
                    row = ""
                    for mounting in parts:
                        if len(parts[mounting]) > i:
                            row += parts[mounting][i] + ","
                        else:
                            row += ","
                    if i == 0:
                        if bluetooth in ModuleParts:
                            row += "Yes,"
                        else:
                            row += "No,"
                    else:
                        row += ","
                    f.write(row + "\n")
else:
    print("Error - Bad parameters")
    print('Eg: py icc3.py iccPart iccPorperty')
