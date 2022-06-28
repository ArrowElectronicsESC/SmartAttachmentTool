import openpyxl

from PyQt5 import QtCore, QtGui, QtWidgets, QtPrintSupport
from PyQt5.QtWidgets import QPushButton, QLineEdit, QLabel


def lowercase_and_remove_non_alphanumeric(s):
    return ''.join(c for c in s if c.isalnum()).lower()


pxl_doc = openpyxl.load_workbook("./Bluetooth Selector Guide.xlsx")
sheet = pxl_doc['Chip Level']

chipLevelParts = set([])
for i in range(1, sheet.max_row+1):
    bt = sheet.cell(row=i, column=3).value
    if bt is not None:
        chipLevelParts.add(lowercase_and_remove_non_alphanumeric(str(bt)))

ModuleParts = set([])
sheet = pxl_doc['Modules']
for i in range(1, sheet.max_row+1):
    bt = sheet.cell(row=i, column=3).value
    if bt is not None:
        ModuleParts.add(lowercase_and_remove_non_alphanumeric(str(bt)))

pxl_doc = openpyxl.load_workbook("./Mod-Antenna relation.xlsx")
sheet = pxl_doc['NORDIC']

bluetooth2antenna = {}

bluetooth = None
for i in range(1, sheet.max_row+1):
    bt = sheet.cell(row=i, column=2).value
    if bt is not None:
        bluetooth = lowercase_and_remove_non_alphanumeric(bt)
        if bluetooth not in bluetooth2antenna:
            bluetooth2antenna[bluetooth] = []
    anthenna = sheet.cell(row=i, column=3).value
    if anthenna is not None:
        bluetooth2antenna[bluetooth].append(anthenna)


pxl_doc = openpyxl.load_workbook("./Mod-Antenna relation.xlsx")
sheet = pxl_doc['TELIT']

bluetooth = None
for i in range(1, sheet.max_row+1):
    bt = sheet.cell(row=i, column=2).value
    if bt is not None:
        bluetooth = lowercase_and_remove_non_alphanumeric(bt)
        if bluetooth not in bluetooth2antenna:
            bluetooth2antenna[bluetooth] = []
    anthenna = sheet.cell(row=i, column=3).value
    if anthenna is not None:
        bluetooth2antenna[bluetooth].append(anthenna)
print(bluetooth2antenna)


# GUI with a text field and a button
class Window(QtWidgets.QWidget):
    def __init__(self):
        super(Window, self).__init__()

        self.setWindowTitle("Smart Attachment Tool")
        self.setGeometry(50, 50, 500, 300)
        # self.setWindowIcon(QtGui.QIcon('pythonlogo.png'))

        self.bluetoothLabel = QLabel("Bluetooth part number:", self)
        self.bluetoothLabel.move(20, 20)

        self.textbox = QLineEdit(self)
        self.textbox.move(20, 40)
        self.textbox.resize(280, 40)

        self.button = QPushButton("Find Attachment", self)
        self.button.move(20, 80)
        self.button.clicked.connect(self.showDialog)

        self.antennaLabel = QLabel(self)
        self.antennaLabel.move(20, 120)
        self.antennaLabel.resize(280, 40)
        self.antennaLabel.setText("Antennas:")

        self.label = QLabel(self)
        self.label.setText("")
        self.label.move(20, 160)
        # self.label.move(20, 120)
        self.label.resize(280, 40)

        self.additionalInfoLabel = QLabel(self)
        self.additionalInfoLabel.move(20, 200)
        self.additionalInfoLabel.resize(280, 40)

        self.show()

    def showDialog(self):
        bluetooth = lowercase_and_remove_non_alphanumeric(self.textbox.text())
        self.additionalInfoLabel.setText("")
        self.label.setText("")
        if bluetooth in bluetooth2antenna:
            if bluetooth in chipLevelParts:
                anthennas = bluetooth2antenna[bluetooth]
                displayText = ""
                for anthenna in anthennas:
                    name = str(anthenna)
                    displayText += name + "\n"
                    print(name)
                self.label.setText(displayText)
                if bluetooth in ModuleParts:
                    self.additionalInfoLabel.setText("There is a module for this part.")
                else:
                    self.additionalInfoLabel.setText("There is no module for this part.")
                # print(bluetooth, ": ", anthenna)
            else:
                self.additionalInfoLabel.setText("Part not found.")
        else:
            self.additionalInfoLabel.setText("Part not found.")


app = QtWidgets.QApplication([])
application = Window()
app.exec_()
